"""استيراد السجل المرجعي للأكواد — قرار مالك 2026-08-02.

يقرأ ملف CHI الرسمي (SBS_V2_Code_list.xlsx — تبويب Technical List المصمم أصلاً
لتغذية أنظمة المعلومات الصحية) أو أي CSV عام، ويُدخل/يحدّث registry_codes بدور المالك
(دور التطبيق SELECT فقط). idempotent: upsert بمفتاح (code_system, code_norm).

الاستخدام:
    python scripts/import_codes.py --file SBS_V2_Code_list.xlsx --system SBS --version "SBS V2.0"
    python scripts/import_codes.py --file icd10am.csv --system ICD10AM --version "ICD-10-AM 12th ed."

CSV العام (ترويسة إلزامية): code, short_desc [, long_desc, chapter, block,
    effective_date, inactive_date, replaced_by]  — التواريخ ISO (YYYY-MM-DD).

تنبيه المصادر: SBS يُنزَّل من chi.gov.sa · ICD-10-AM مرخّص (IHACPA) عبر قنوات CHI/nphies.
**لا يُستورد ICD-10-CM الأمريكي ولا DRG — nphies يعتمد ICD-10-AM وSBS حصراً.**
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import re
import sys
from pathlib import Path
from typing import Any, Iterator

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "backend"))

from sqlalchemy import create_engine, select  # noqa: E402
from sqlalchemy.orm import Session  # noqa: E402

from app.config import get_settings  # noqa: E402
from app.models import RegistryCode  # noqa: E402
from app.services.code_registry import normalize_code  # noqa: E402


def _to_date(value: Any) -> dt.date | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()[:10]
    try:
        return dt.date.fromisoformat(text)
    except ValueError:
        return None


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def _header_index(header: list[Any], *needles: str) -> int | None:
    """يطابق عمود الترويسة باحتواء غير حساس للحالة/الفراغات — ترويسات CHI فيها فراغات زائدة."""
    for index, cell in enumerate(header):
        name = re.sub(r"\s+", " ", str(cell or "")).strip().lower()
        if name and all(needle in name for needle in needles):
            return index
    return None


def rows_from_chi_xlsx(path: Path, sheet: str | None) -> Iterator[dict[str, Any]]:
    """تبويب «Technical List» من ملف CHI — الأعمدة تُكتشف بالترويسة لا بالموضع."""
    try:
        import openpyxl
    except ImportError:  # openpyxl ليس تبعية للباك اند — يلزم للاستيراد فقط
        raise SystemExit("openpyxl غير مثبت — نفّذ: pip install openpyxl")

    workbook = openpyxl.load_workbook(path, read_only=True)
    sheet_name = sheet
    if sheet_name is None:
        candidates = [ws for ws in workbook.sheetnames if "technical" in ws.lower()]
        if not candidates:
            raise SystemExit(f"لا تبويب Technical List في {path.name} — مرّر --sheet صراحة")
        sheet_name = candidates[0]
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    header = list(next(rows))

    col_code = _header_index(header, "code", "hyphen")  # الصيغة القانونية 40803-00-00
    col_code_plain = _header_index(header, "sbs code") if col_code is None else None
    col_short = _header_index(header, "short desc")
    col_long = _header_index(header, "long desc")
    col_effective = _header_index(header, "effective date")
    col_inactive = _header_index(header, "inactive date")
    col_mapping = _header_index(header, "inactive code mapping")
    col_chapter = _header_index(header, "chapter name")
    col_block = _header_index(header, "block name")
    if (col_code is None and col_code_plain is None) or col_short is None:
        raise SystemExit(f"ترويسة غير متوقعة في تبويب {sheet_name}: {header}")

    for row in rows:
        code = _clean(row[col_code] if col_code is not None else row[col_code_plain])
        if not code:
            continue
        inactive_date = _to_date(row[col_inactive]) if col_inactive is not None else None
        replaced_by = _clean(row[col_mapping]) if col_mapping is not None else None
        if replaced_by:
            # «Replaced with  42845-01-01» → «42845-01-01»
            match = re.search(r"[\dA-Z][\dA-Z\-.]+$", replaced_by)
            replaced_by = match.group(0) if match else replaced_by
        yield {
            "code": code,
            "short_desc": _clean(row[col_short]) or code,
            "long_desc": _clean(row[col_long]) if col_long is not None else None,
            "chapter": _clean(row[col_chapter]) if col_chapter is not None else None,
            "block": _clean(row[col_block]) if col_block is not None else None,
            "effective_date": _to_date(row[col_effective]) if col_effective is not None else None,
            "inactive_date": inactive_date,
            "is_active": inactive_date is None,
            "replaced_by": replaced_by,
        }


def rows_from_csv(path: Path) -> Iterator[dict[str, Any]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            code = _clean(row.get("code"))
            if not code:
                continue
            inactive_date = _to_date(row.get("inactive_date"))
            yield {
                "code": code,
                "short_desc": _clean(row.get("short_desc")) or code,
                "long_desc": _clean(row.get("long_desc")),
                "chapter": _clean(row.get("chapter")),
                "block": _clean(row.get("block")),
                "effective_date": _to_date(row.get("effective_date")),
                "inactive_date": inactive_date,
                "is_active": inactive_date is None,
                "replaced_by": _clean(row.get("replaced_by")),
            }


def import_codes(db: Session, system: str, version: str, rows: Iterator[dict[str, Any]]) -> tuple[int, int]:
    existing = {
        row.code_norm: row
        for row in db.execute(select(RegistryCode).where(RegistryCode.code_system == system)).scalars()
    }
    inserted = updated = 0
    for data in rows:
        norm = normalize_code(data["code"])
        entry = existing.get(norm)
        if entry is None:
            entry = RegistryCode(code_system=system, code_norm=norm, registry_version=version, **data)
            db.add(entry)
            existing[norm] = entry
            inserted += 1
        else:
            for field, value in {**data, "registry_version": version}.items():
                setattr(entry, field, value)
            updated += 1
    return inserted, updated


def main() -> None:
    parser = argparse.ArgumentParser(description="استيراد السجل المرجعي للأكواد")
    parser.add_argument("--file", required=True, help="ملف xlsx (CHI) أو csv عام")
    parser.add_argument("--system", required=True, choices=["SBS", "ICD10AM", "ACHI", "SFDA", "GMDN"])
    parser.add_argument("--version", required=True, help='إصدار السجل، مثل "SBS V2.0"')
    parser.add_argument("--sheet", default=None, help="اسم تبويب xlsx (الافتراضي: اكتشاف Technical List)")
    parser.add_argument("--dry-run", action="store_true", help="عدّ بلا كتابة")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        raise SystemExit(f"الملف غير موجود: {path}")
    rows = rows_from_chi_xlsx(path, args.sheet) if path.suffix.lower() in (".xlsx", ".xlsm") else rows_from_csv(path)

    settings = get_settings()
    url = os.environ.get("MIGRATIONS_DATABASE_URL") or settings.migrations_database_url or settings.database_url
    engine = create_engine(url)
    with Session(engine) as db:
        inserted, updated = import_codes(db, args.system, args.version, rows)
        if args.dry_run:
            db.rollback()
            print(f"dry-run: كان سيُدخل {inserted} ويحدّث {updated} — {args.system} {args.version}")
        else:
            db.commit()
            print(f"اكتمل: أُدخل {inserted} وحُدّث {updated} كوداً — {args.system} {args.version}")


if __name__ == "__main__":
    main()
