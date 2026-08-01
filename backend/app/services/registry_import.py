"""استيراد السجل المرجعي للأكواد — المنطق المشترك بين CLI (scripts/import_codes.py)
وكونسول السوبر أدمن (POST /sa/registry/import) — قرار مالك 2026-08-02.

يقرأ ملف CHI الرسمي (تبويب «Technical List» بترويسته) أو CSV عاماً، ويكتب upsert
بمفتاح (code_system, code_norm). ملف غير مفهوم يرفع ValueError — يترجمه كل مدخل لخطئه.
"""
from __future__ import annotations

import csv
import datetime as dt
import io
import re
from typing import Any, BinaryIO, Iterator

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from ..models import RegistryCode
from .code_registry import normalize_code

REGISTRY_SYSTEMS = ("ICD10AM", "ACHI", "SBS", "SFDA", "GMDN")


def _to_date(value: Any) -> dt.date | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()[:10]
    try:
        return dt.date.fromisoformat(text)
    except ValueError:
        return None


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def _header_index(header: list[Any], *needles: str) -> int | None:
    """يطابق عمود الترويسة باحتواء غير حساس للحالة/الفراغات — ترويسات CHI فيها فراغات زائدة."""
    for index, cell in enumerate(header):
        name = re.sub(r"\s+", " ", str(cell or "")).strip().lower()
        if name and all(needle in name for needle in needles):
            return index
    return None


def rows_from_chi_xlsx(stream: BinaryIO, sheet: str | None = None) -> Iterator[dict[str, Any]]:
    """تبويب «Technical List» من ملف CHI — الأعمدة تُكتشف بالترويسة لا بالموضع."""
    try:
        import openpyxl
    except ImportError as exc:  # تبعية معلنة في pyproject — حارس لبيئات ناقصة
        raise ValueError("openpyxl غير مثبت — لا يمكن قراءة xlsx") from exc

    try:
        workbook = openpyxl.load_workbook(stream, read_only=True)
    except Exception as exc:
        raise ValueError(f"ملف xlsx غير صالح: {exc}") from exc
    sheet_name = sheet
    if sheet_name is None:
        candidates = [ws for ws in workbook.sheetnames if "technical" in ws.lower()]
        if not candidates:
            raise ValueError("لا تبويب Technical List في الملف — حدد التبويب صراحة")
        sheet_name = candidates[0]
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"التبويب غير موجود: {sheet_name}")
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration as exc:
        raise ValueError("التبويب فارغ") from exc

    col_code = _header_index(header, "code", "hyphen")  # الصيغة القانونية 40803-00-00
    col_code_plain = _header_index(header, "sbs code") if col_code is None else None
    col_short = _header_index(header, "short desc")
    col_long = _header_index(header, "long desc")
    col_effective = _header_index(header, "effective date")
    col_inactive = _header_index(header, "inactive date")
    col_mapping = _header_index(header, "inactive code mapping")
    col_chapter = _header_index(header, "chapter name")
    col_block = _header_index(header, "block name")
    if (col_code is None and col_code_plain is None) or col_short is None:
        raise ValueError(f"ترويسة غير متوقعة في تبويب {sheet_name}: {header}")

    for row in rows:
        code = _clean(row[col_code] if col_code is not None else row[col_code_plain])
        if not code:
            continue
        inactive_date = _to_date(row[col_inactive]) if col_inactive is not None else None
        replaced_by = _clean(row[col_mapping]) if col_mapping is not None else None
        if replaced_by:
            # «Replaced with  42845-01-01» → «42845-01-01»
            match = re.search(r"[\dA-Z][\dA-Z\-.]+$", replaced_by)
            replaced_by = match.group(0) if match else replaced_by
        yield {
            "code": code,
            "short_desc": _clean(row[col_short]) or code,
            "long_desc": _clean(row[col_long]) if col_long is not None else None,
            "chapter": _clean(row[col_chapter]) if col_chapter is not None else None,
            "block": _clean(row[col_block]) if col_block is not None else None,
            "effective_date": _to_date(row[col_effective]) if col_effective is not None else None,
            "inactive_date": inactive_date,
            "is_active": inactive_date is None,
            "replaced_by": replaced_by,
        }


def rows_from_csv_bytes(data: bytes) -> Iterator[dict[str, Any]]:
    """CSV عام بترويسة: code, short_desc [, long_desc, chapter, block,
    effective_date, inactive_date, replaced_by] — التواريخ ISO."""
    try:
        text_stream = io.StringIO(data.decode("utf-8-sig"))
    except UnicodeDecodeError as exc:
        raise ValueError("ترميز CSV غير مقروء — يلزم UTF-8") from exc
    reader = csv.DictReader(text_stream)
    if not reader.fieldnames or "code" not in [f.strip().lower() for f in reader.fieldnames]:
        raise ValueError(f"ترويسة CSV بلا عمود code: {reader.fieldnames}")
    for row in reader:
        row = {(k or "").strip().lower(): v for k, v in row.items()}
        code = _clean(row.get("code"))
        if not code:
            continue
        inactive_date = _to_date(row.get("inactive_date"))
        yield {
            "code": code,
            "short_desc": _clean(row.get("short_desc")) or code,
            "long_desc": _clean(row.get("long_desc")),
            "chapter": _clean(row.get("chapter")),
            "block": _clean(row.get("block")),
            "effective_date": _to_date(row.get("effective_date")),
            "inactive_date": inactive_date,
            "is_active": inactive_date is None,
            "replaced_by": _clean(row.get("replaced_by")),
        }


def parse_registry_file(filename: str, content: bytes, sheet: str | None = None) -> Iterator[dict[str, Any]]:
    """يختار المحلل من الامتداد — xlsx (CHI) أو csv عام."""
    lowered = (filename or "").lower()
    if lowered.endswith((".xlsx", ".xlsm")):
        return rows_from_chi_xlsx(io.BytesIO(content), sheet)
    if lowered.endswith(".csv"):
        return rows_from_csv_bytes(content)
    raise ValueError(f"صيغة غير مدعومة: {filename} — المقبول xlsx (CHI) أو csv")


def import_codes(
    db: Session, system: str, version: str, rows: Iterator[dict[str, Any]]
) -> tuple[int, int]:
    """upsert بمفتاح (code_system, code_norm) — يعيد (المُدخل، المُحدَّث)."""
    existing = {
        row.code_norm: row
        for row in db.execute(select(RegistryCode).where(RegistryCode.code_system == system)).scalars()
    }
    inserted = updated = 0
    for data in rows:
        norm = normalize_code(data["code"])
        entry = existing.get(norm)
        if entry is None:
            entry = RegistryCode(code_system=system, code_norm=norm, registry_version=version, **data)
            db.add(entry)
            existing[norm] = entry
            inserted += 1
        else:
            for field, value in {**data, "registry_version": version}.items():
                setattr(entry, field, value)
            updated += 1
    if inserted == 0 and updated == 0:
        raise ValueError("الملف بلا صفوف أكواد صالحة")
    return inserted, updated


def registry_overview(db: Session) -> list[dict[str, Any]]:
    """حالة كل نظام: الأعداد والإصدارات وآخر تحديث — للعرض في W-SA وحالة الإنفاذ."""
    stats = {
        row.code_system: row
        for row in db.execute(
            select(
                RegistryCode.code_system,
                func.count(RegistryCode.id).label("total"),
                func.count(RegistryCode.id).filter(RegistryCode.is_active).label("active"),
                func.max(RegistryCode.updated_at).label("last_updated"),
                func.array_agg(RegistryCode.registry_version.distinct()).label("versions"),
            ).group_by(RegistryCode.code_system)
        )
    }
    out: list[dict[str, Any]] = []
    for system in REGISTRY_SYSTEMS:
        row = stats.get(system)
        out.append({
            "system": system,
            "total": row.total if row else 0,
            "active": row.active if row else 0,
            "inactive": (row.total - row.active) if row else 0,
            "versions": sorted(row.versions) if row else [],
            "last_updated": row.last_updated.isoformat() if row else None,
            "enforced": bool(row and row.total),  # سجل فارغ = لا تحقق لهذا النظام
        })
    return out
